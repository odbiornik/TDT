from __future__ import annotations

import csv
import json
import math
from html import escape
from pathlib import Path
from typing import Any

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
except ImportError:  # pragma: no cover - optional runtime dependency
    go = None
    make_subplots = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover - optional runtime dependency
    Workbook = None
    Font = None
    PatternFill = None
    get_column_letter = None


def find_project_root(start: Path | None = None) -> Path:
    base = (start or Path.cwd()).resolve()
    candidates = [base, *base.parents]
    for candidate in candidates:
        if (candidate / "data" / "sessions").exists():
            return candidate
        if candidate.name == "src" and (candidate.parent / "data" / "sessions").exists():
            return candidate.parent
    raise FileNotFoundError("Could not find the project root containing data/sessions.")


def latest_session_dir(project_root: Path | None = None) -> Path:
    root = project_root or find_project_root()
    sessions_dir = root / "data" / "sessions"
    return max((path for path in sessions_dir.iterdir() if path.is_dir()), key=lambda path: path.stat().st_mtime)


def parse_number(value: Any, cast=float) -> Any:
    if value in (None, ""):
        return None
    return cast(value)


def load_session(session_dir: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    summary_path = session_dir / "summary.json"
    trials_path = session_dir / "trials.csv"

    with summary_path.open("r", encoding="utf-8") as handle:
        session_summary = json.load(handle)

    with trials_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))

    for index, row in enumerate(rows, start=1):
        row["session_trial_number"] = parse_number(row.get("session_trial_number"), int) or index
        row["attempt_number"] = parse_number(row.get("attempt_number"), int)
        row["requested_soa_ms"] = parse_number(row.get("requested_soa_ms"))
        row["rt_ms"] = parse_number(row.get("rt_ms"))
        row["threshold50_ms"] = parse_number(row.get("threshold50_ms"))
        row["threshold75_ms"] = parse_number(row.get("threshold75_ms"))
        row["guess_rate"] = parse_number(row.get("guess_rate"))
        row["ci50_width_ms"] = parse_number(row.get("ci50_width_ms"))
        row["ci50_low_ms"] = parse_number(row.get("ci50_low_ms"))
        row["ci50_high_ms"] = parse_number(row.get("ci50_high_ms"))
        row["ci75_width_ms"] = parse_number(row.get("ci75_width_ms"))
        row["ci75_low_ms"] = parse_number(row.get("ci75_low_ms"))
        row["ci75_high_ms"] = parse_number(row.get("ci75_high_ms"))
        row["entropy_normalized"] = parse_number(row.get("entropy_normalized"))
        sanity_correct = row.get("sanity_correct")
        row["sanity_correct"] = (
            None if sanity_correct in (None, "") else str(sanity_correct).lower() == "true"
        )

    for row in rows:
        row["display_trial_number"] = row["session_trial_number"]
        row["trial_local_number"] = row["attempt_number"]
        if row.get("threshold50_ms") is not None and row.get("ci50_width_ms") is not None:
            if row.get("ci50_low_ms") is None:
                row["ci50_low_ms"] = row["threshold50_ms"] - (row["ci50_width_ms"] / 2.0)
            if row.get("ci50_high_ms") is None:
                row["ci50_high_ms"] = row["threshold50_ms"] + (row["ci50_width_ms"] / 2.0)
        if row.get("threshold75_ms") is not None:
            if row.get("ci75_width_ms") is None and row.get("ci50_width_ms") is not None:
                row["ci75_width_ms"] = row["ci50_width_ms"]
            if (
                row.get("ci75_low_ms") is None
                and row.get("ci50_low_ms") is not None
                and row.get("threshold50_ms") is not None
            ):
                row["ci75_low_ms"] = row["ci50_low_ms"] + (row["threshold75_ms"] - row["threshold50_ms"])
            if (
                row.get("ci75_high_ms") is None
                and row.get("ci50_high_ms") is not None
                and row.get("threshold50_ms") is not None
            ):
                row["ci75_high_ms"] = row["ci50_high_ms"] + (row["threshold75_ms"] - row["threshold50_ms"])
            if row.get("ci75_low_ms") is None and row.get("ci75_width_ms") is not None:
                row["ci75_low_ms"] = row["threshold75_ms"] - (row["ci75_width_ms"] / 2.0)
            if row.get("ci75_high_ms") is None and row.get("ci75_width_ms") is not None:
                row["ci75_high_ms"] = row["threshold75_ms"] + (row["ci75_width_ms"] / 2.0)

    return session_summary, rows


def _require_plotly() -> None:
    if go is None or make_subplots is None:
        raise RuntimeError("Plotly is not installed. Install plotly to export session plots.")


def _nice_tick_step(span: float) -> float:
    span = max(float(span), 1.0)
    raw_step = span / 7.0
    magnitude = 10 ** math.floor(math.log10(raw_step))
    residual = raw_step / magnitude
    if residual <= 1:
        factor = 1
    elif residual <= 2:
        factor = 2
    elif residual <= 5:
        factor = 5
    else:
        factor = 10
    return float(factor * magnitude)


def build_session_figure(session_dir: Path, session_summary: dict[str, Any], rows: list[dict[str, Any]]):
    _require_plotly()

    practice_rows = [row for row in rows if row.get("phase") == "practice"]
    main_rows = [row for row in rows if row.get("phase") == "main"]
    sanity_rows = [row for row in rows if row.get("phase") == "sanity"]
    updated_main_rows = [row for row in main_rows if row.get("updated_staircase") == "True"]

    practice_yes_rows = [row for row in practice_rows if row.get("response") == "yes"]
    practice_no_rows = [row for row in practice_rows if row.get("response") == "no"]
    yes_rows = [row for row in main_rows if row.get("response") == "yes"]
    no_rows = [row for row in main_rows if row.get("response") == "no"]

    sanity_zero_rows = [row for row in sanity_rows if row.get("sanity_kind") == "zero"]
    sanity_easy_rows = [row for row in sanity_rows if row.get("sanity_kind") == "easy"]
    sanity_zero_yes_rows = [row for row in sanity_zero_rows if row.get("response") == "yes"]
    sanity_zero_no_rows = [row for row in sanity_zero_rows if row.get("response") == "no"]
    sanity_easy_yes_rows = [row for row in sanity_easy_rows if row.get("response") == "yes"]
    sanity_easy_no_rows = [row for row in sanity_easy_rows if row.get("response") == "no"]

    summary_block = session_summary["summary"]
    estimates = summary_block["estimates"]
    reliability = summary_block["reliability"]

    fig = make_subplots(
        rows=1,
        cols=1,
        specs=[[{"secondary_y": True}]],
        subplot_titles=("Trial sequence and threshold estimate",),
    )

    if practice_rows:
        fig.add_trace(
            go.Scatter(
                x=[row["display_trial_number"] for row in practice_yes_rows],
                y=[row["requested_soa_ms"] for row in practice_yes_rows],
                mode="markers",
                name="Practice Yes",
                showlegend=False,
                customdata=[row["trial_local_number"] for row in practice_yes_rows],
                marker=dict(color="rgba(31, 119, 180, 0.70)", size=8, symbol="circle-open"),
                hovertemplate="Practice trial %{customdata}<br>Interval=%{y} ms<extra>Yes</extra>",
            ),
            row=1,
            col=1,
            secondary_y=False,
        )
        fig.add_trace(
            go.Scatter(
                x=[row["display_trial_number"] for row in practice_no_rows],
                y=[row["requested_soa_ms"] for row in practice_no_rows],
                mode="markers",
                name="Practice No",
                showlegend=False,
                customdata=[row["trial_local_number"] for row in practice_no_rows],
                marker=dict(color="rgba(214, 39, 40, 0.70)", size=8, symbol="circle-open"),
                hovertemplate="Practice trial %{customdata}<br>Interval=%{y} ms<extra>No</extra>",
            ),
            row=1,
            col=1,
            secondary_y=False,
        )

    fig.add_trace(
        go.Scatter(
            x=[row["display_trial_number"] for row in yes_rows],
            y=[row["requested_soa_ms"] for row in yes_rows],
            mode="markers",
            name="Yes",
            legendgroup="responses",
            legendgrouptitle_text="Responses",
            customdata=[row["trial_local_number"] for row in yes_rows],
            marker=dict(color="#1f77b4", size=8, symbol="circle"),
            hovertemplate="Main trial %{customdata}<br>Interval=%{y} ms<extra>Yes</extra>",
        ),
        row=1,
        col=1,
        secondary_y=False,
    )

    for sanity_rows_subset, trace_name, trace_color, trace_symbol in [
        (sanity_zero_yes_rows, "0 ms", "#1f77b4", "square-open"),
        (sanity_zero_no_rows, "0 ms", "#d62728", "square-open"),
        (sanity_easy_yes_rows, "High interval", "#1f77b4", "diamond-open"),
        (sanity_easy_no_rows, "High interval", "#d62728", "diamond-open"),
    ]:
        fig.add_trace(
            go.Scatter(
                x=[row["display_trial_number"] for row in sanity_rows_subset],
                y=[row["requested_soa_ms"] for row in sanity_rows_subset],
                mode="markers",
                name=trace_name,
                legendgroup="checks",
                showlegend=False,
                customdata=[
                    [
                        row["trial_local_number"],
                        row.get("response", ""),
                        row.get("expected_response", ""),
                        row.get("sanity_correct"),
                    ]
                    for row in sanity_rows_subset
                ],
                marker=dict(color=trace_color, size=10, symbol=trace_symbol),
                hovertemplate="Sanity trial %{customdata[0]}<br>Interval=%{y} ms<br>Response=%{customdata[1]}<br>Expected=%{customdata[2]}<br>Correct=%{customdata[3]}<extra></extra>",
            ),
            row=1,
            col=1,
            secondary_y=False,
        )

    fig.add_trace(
        go.Scatter(
            x=[None],
            y=[None],
            mode="markers",
            name="0 ms",
            legendgroup="checks",
            legendgrouptitle_text="Sanity checks",
            marker=dict(color="rgba(120, 120, 120, 0.85)", size=10, symbol="square-open"),
            hoverinfo="skip",
        ),
        row=1,
        col=1,
        secondary_y=False,
    )

    fig.add_trace(
        go.Scatter(
            x=[None],
            y=[None],
            mode="markers",
            name="High interval",
            legendgroup="checks",
            marker=dict(color="rgba(120, 120, 120, 0.85)", size=10, symbol="diamond-open"),
            hoverinfo="skip",
        ),
        row=1,
        col=1,
        secondary_y=False,
    )

    fig.add_trace(
        go.Scatter(
            x=[row["display_trial_number"] for row in no_rows],
            y=[row["requested_soa_ms"] for row in no_rows],
            mode="markers",
            name="No",
            legendgroup="responses",
            customdata=[row["trial_local_number"] for row in no_rows],
            marker=dict(color="#d62728", size=8, symbol="circle"),
            hovertemplate="Main trial %{customdata}<br>Interval=%{y} ms<extra>No</extra>",
        ),
        row=1,
        col=1,
        secondary_y=False,
    )

    ci50_band_rows = [
        row for row in updated_main_rows if row.get("ci50_low_ms") is not None and row.get("ci50_high_ms") is not None
    ]
    ci75_band_rows = [
        row for row in updated_main_rows if row.get("ci75_low_ms") is not None and row.get("ci75_high_ms") is not None
    ]

    if ci50_band_rows:
        fig.add_trace(
            go.Scatter(
                x=[row["display_trial_number"] for row in ci50_band_rows],
                y=[row["ci50_low_ms"] for row in ci50_band_rows],
                customdata=[
                    [row["trial_local_number"], row["requested_soa_ms"], row["ci50_low_ms"], row["ci50_high_ms"]]
                    for row in ci50_band_rows
                ],
                mode="lines",
                name="T50 CI range",
                legendgroup="threshold50",
                showlegend=False,
                line=dict(color="rgba(255, 127, 14, 0.35)", width=1, dash="dot"),
                hovertemplate="Main trial %{customdata[0]}<br>Interval=%{customdata[1]} ms<br>T50 CI range=[%{customdata[2]:.2f}, %{customdata[3]:.2f}] ms<extra>CI lower</extra>",
            ),
            row=1,
            col=1,
            secondary_y=True,
        )
        fig.add_trace(
            go.Scatter(
                x=[row["display_trial_number"] for row in ci50_band_rows],
                y=[row["ci50_high_ms"] for row in ci50_band_rows],
                customdata=[
                    [row["trial_local_number"], row["requested_soa_ms"], row["ci50_low_ms"], row["ci50_high_ms"]]
                    for row in ci50_band_rows
                ],
                mode="lines",
                name="T50 CI range",
                legendgroup="threshold50",
                legendgrouptitle_text="Threshold 50%",
                showlegend=True,
                fill="tonexty",
                fillcolor="rgba(255, 127, 14, 0.15)",
                line=dict(color="rgba(255, 127, 14, 0.35)", width=1, dash="dot"),
                hovertemplate="Main trial %{customdata[0]}<br>Interval=%{customdata[1]} ms<br>T50 CI range=[%{customdata[2]:.2f}, %{customdata[3]:.2f}] ms<extra>CI range</extra>",
            ),
            row=1,
            col=1,
            secondary_y=True,
        )

    fig.add_trace(
        go.Scatter(
            x=[row["display_trial_number"] for row in updated_main_rows],
            y=[row["threshold50_ms"] for row in updated_main_rows],
            customdata=[
                [row["trial_local_number"], row["requested_soa_ms"], row.get("ci50_low_ms"), row.get("ci50_high_ms")]
                for row in updated_main_rows
            ],
            mode="lines+markers",
            name="Threshold 50%",
            legendgroup="threshold50",
            legendgrouptitle_text=(None if ci50_band_rows else "Threshold 50%"),
            line=dict(color="#111111", width=2),
            marker=dict(size=5),
            hovertemplate="Main trial %{customdata[0]}<br>Interval=%{customdata[1]} ms<br>T50=%{y:.2f} ms<br>T50 CI range=[%{customdata[2]:.2f}, %{customdata[3]:.2f}] ms<extra></extra>",
        ),
        row=1,
        col=1,
        secondary_y=True,
    )

    if ci75_band_rows:
        fig.add_trace(
            go.Scatter(
                x=[row["display_trial_number"] for row in ci75_band_rows],
                y=[row["ci75_low_ms"] for row in ci75_band_rows],
                customdata=[
                    [row["trial_local_number"], row["requested_soa_ms"], row["ci75_low_ms"], row["ci75_high_ms"]]
                    for row in ci75_band_rows
                ],
                mode="lines",
                name="T75 CI range",
                legendgroup="threshold75",
                showlegend=False,
                line=dict(color="rgba(128, 128, 128, 0.35)", width=1, dash="dot"),
                hovertemplate="Main trial %{customdata[0]}<br>Interval=%{customdata[1]} ms<br>T75 CI range=[%{customdata[2]:.2f}, %{customdata[3]:.2f}] ms<extra>CI lower</extra>",
            ),
            row=1,
            col=1,
            secondary_y=True,
        )
        fig.add_trace(
            go.Scatter(
                x=[row["display_trial_number"] for row in ci75_band_rows],
                y=[row["ci75_high_ms"] for row in ci75_band_rows],
                customdata=[
                    [row["trial_local_number"], row["requested_soa_ms"], row["ci75_low_ms"], row["ci75_high_ms"]]
                    for row in ci75_band_rows
                ],
                mode="lines",
                name="T75 CI range",
                legendgroup="threshold75",
                legendgrouptitle_text="Threshold 75%",
                showlegend=True,
                fill="tonexty",
                fillcolor="rgba(128, 128, 128, 0.12)",
                line=dict(color="rgba(128, 128, 128, 0.35)", width=1, dash="dot"),
                hovertemplate="Main trial %{customdata[0]}<br>Interval=%{customdata[1]} ms<br>T75 CI range=[%{customdata[2]:.2f}, %{customdata[3]:.2f}] ms<extra>CI range</extra>",
            ),
            row=1,
            col=1,
            secondary_y=True,
        )

    fig.add_trace(
        go.Scatter(
            x=[row["display_trial_number"] for row in updated_main_rows],
            y=[row["threshold75_ms"] for row in updated_main_rows],
            customdata=[
                [row["trial_local_number"], row["requested_soa_ms"], row.get("ci75_low_ms"), row.get("ci75_high_ms")]
                for row in updated_main_rows
            ],
            mode="lines+markers",
            name="Threshold 75%",
            legendgroup="threshold75",
            legendgrouptitle_text=(None if ci75_band_rows else "Threshold 75%"),
            line=dict(color="#808080", width=1.8),
            marker=dict(size=4),
            hovertemplate="Main trial %{customdata[0]}<br>Interval=%{customdata[1]} ms<br>T75=%{y:.2f} ms<br>T75 CI range=[%{customdata[2]:.2f}, %{customdata[3]:.2f}] ms<extra></extra>",
        ),
        row=1,
        col=1,
        secondary_y=True,
    )

    y_values = [
        float(row["requested_soa_ms"])
        for row in (practice_rows + main_rows + sanity_rows)
        if row.get("requested_soa_ms") is not None
    ]
    for row in updated_main_rows:
        for key in ("threshold50_ms", "threshold75_ms"):
            if row.get(key) is not None:
                y_values.append(float(row[key]))
    for row in ci50_band_rows:
        y_values.extend([float(row["ci50_low_ms"]), float(row["ci50_high_ms"])])
    for row in ci75_band_rows:
        y_values.extend([float(row["ci75_low_ms"]), float(row["ci75_high_ms"])])

    if not y_values:
        y_values = [0.0, 100.0]

    y_min = min(y_values)
    y_max = max(y_values)
    if math.isclose(y_min, y_max):
        y_max = y_min + 1.0
    span = y_max - y_min
    dtick = _nice_tick_step(span)
    bottom_padding = max(2.5, dtick * 0.35)
    top_padding = max(dtick * 0.5, span * 0.05)
    y_axis_min = float(y_min - bottom_padding)
    y_axis_max = math.ceil((y_max + top_padding) / dtick) * dtick

    if practice_rows:
        fig.add_vline(
            x=max(row["display_trial_number"] for row in practice_rows) + 0.5,
            line_dash="dot",
            line_color="gray",
            row=1,
            col=1,
        )

    fig.update_yaxes(
        title_text="Stimulus interval (ms)",
        row=1,
        col=1,
        secondary_y=False,
        range=[y_axis_min, y_axis_max],
        tickmode="linear",
        tick0=0.0,
        dtick=dtick,
    )
    fig.update_yaxes(
        title_text="Threshold estimate (ms)",
        row=1,
        col=1,
        secondary_y=True,
        range=[y_axis_min, y_axis_max],
        tickmode="linear",
        tick0=0.0,
        dtick=dtick,
        matches="y",
        showgrid=False,
    )

    tickvals = []
    ticktext = []
    if practice_rows:
        tickvals.append((practice_rows[0]["display_trial_number"] + practice_rows[-1]["display_trial_number"]) / 2.0)
        ticktext.append("Practice")
    tickvals.extend(
        row["display_trial_number"] for row in main_rows if row["trial_local_number"] == 1 or row["trial_local_number"] % 10 == 0
    )
    ticktext.extend(
        str(row["trial_local_number"]) for row in main_rows if row["trial_local_number"] == 1 or row["trial_local_number"] % 10 == 0
    )
    fig.update_xaxes(
        title_text="Trial number",
        tickmode="array",
        tickvals=tickvals,
        ticktext=ticktext,
        row=1,
        col=1,
    )
    fig.update_layout(
        height=780,
        template="plotly_white",
        title=dict(
            text=(
                f"IRL session: {session_dir.name}<br>"
                f"T50={estimates['threshold50_ms']:.1f} ms | "
                f"CI range=[{reliability['threshold50_ci95_low_ms']:.2f}, {reliability['threshold50_ci95_high_ms']:.2f}] ms"
            ),
            font=dict(size=16),
            y=0.935,
            yanchor="top",
            x=0.5,
            xanchor="center",
            pad=dict(t=10, b=22),
        ),
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.045,
            xanchor="left",
            x=0,
            traceorder="grouped",
            groupclick="togglegroup",
            entrywidthmode="pixels",
            entrywidth=104,
            tracegroupgap=28,
        ),
        margin=dict(t=235, b=80, l=75, r=75),
    )
    return fig

def build_session_summary_lines(session_summary: dict[str, Any]) -> list[str]:
    summary_block = session_summary["summary"]
    estimates = summary_block["estimates"]
    reliability = summary_block["reliability"]
    guess_rate_value = estimates.get("guess_rate", estimates.get("lapse_rate"))
    stop_checks = summary_block["early_stop"]["checks"]

    entropy_display = (
        "off"
        if session_summary["config"]["stop_criteria"].get("mean_entropy_normalized") in (None, "")
        else str(stop_checks["entropy"])
    )
    sanity_gate_value = session_summary["config"]["stop_criteria"].get("minimum_sanity_correct_rate")
    sanity_display = "off" if sanity_gate_value in (None, "") else str(stop_checks.get("sanity"))
    timing_block = session_summary.get("session_timing", {})

    lines = [
        "Session summary:",
        f"  Trials completed: {summary_block['completed_trials']}/{session_summary['config']['trial_count']}",
        f"  Sanity trials inserted: {session_summary.get('sanity_trials_completed', 0)}",
        f"  Flash brightness: {session_summary['config'].get('flash_rgb_level', 'n/a')}/255",
        f"  Threshold 50%: {estimates['threshold50_ms']:.1f} ms",
        f"  Threshold 75%: {estimates['threshold75_ms']:.1f} ms",
        f"  JND proxy: {estimates['jnd_ms']:.1f} ms",
        f"  Guess rate: {guess_rate_value:.1%}",
        f"  Lapse rate: {estimates['lapse_rate']:.1%}",
        (
            "  Threshold 50% CI95: "
            f"[{reliability['threshold50_ci95_low_ms']:.2f}, {reliability['threshold50_ci95_high_ms']:.2f}] ms"
        ),
        f"  Threshold 50% CI95 width: {reliability['threshold50_ci95_width_ms']:.2f} ms",
        (
            "  Posterior mass within "
            f"+/- {session_summary['config']['stop_criteria']['threshold50_ci95_width_ms'] / 2.0:.1f} ms: "
            f"{reliability['threshold50_mass_within_precision_window']:.1%}"
        ),
        f"  Normalized entropy: {reliability['mean_entropy_normalized']:.3f}",
        (
            "  Early stop checks: "
            f"trials={stop_checks['enough_trials']}, ci50={stop_checks['threshold50']}, "
            f"entropy={entropy_display}, sanity={sanity_display}, active={summary_block['early_stop']['active']}"
        ),
    ]

    if timing_block:
        lines.extend(
            [
                (
                    "  Total time: "
                    f"{timing_block.get('total_duration_display', 'n/a')}"
                    + (
                        ""
                        if timing_block.get("total_duration_s") is None
                        else f" ({timing_block['total_duration_s']:.1f} s)"
                    )
                ),
                (
                    "  Practice duration: "
                    + (
                        "n/a"
                        if timing_block.get("practice_duration_s") in (None, "")
                        else f"{timing_block.get('practice_duration_display', 'n/a')} ({timing_block['practice_duration_s']:.1f} s)"
                    )
                ),
                (
                    "  Adaptive duration: "
                    + (
                        "n/a"
                        if timing_block.get("adaptive_duration_s") in (None, "")
                        else f"{timing_block.get('adaptive_duration_display', 'n/a')} ({timing_block['adaptive_duration_s']:.1f} s)"
                    )
                ),
            ]
        )

    sanity_summary = summary_block.get("sanity_checks")
    if sanity_summary:
        zero_control = sanity_summary["zero_control"]
        easy_control = sanity_summary["easy_control"]
        zero_rate = "n/a" if zero_control["correct_rate"] is None else f"{zero_control['correct_rate']:.0%}"
        easy_rate = "n/a" if easy_control["correct_rate"] is None else f"{easy_control['correct_rate']:.0%}"
        total_rate = "n/a" if sanity_summary["correct_rate_total"] is None else f"{sanity_summary['correct_rate_total']:.0%}"
        lines.extend(
            [
                "  Sanity checks:",
                (
                    "    0 ms controls (expected No): "
                    f"{zero_control['correct_trials']}/{zero_control['valid_trials']} correct ({zero_rate})"
                ),
                (
                    "    High-interval controls (expected Yes): "
                    f"{easy_control['correct_trials']}/{easy_control['valid_trials']} correct ({easy_rate})"
                ),
                (
                    "    Overall sanity accuracy: "
                    f"{sanity_summary['correct_trials_total']}/{sanity_summary['valid_trials_total']} correct ({total_rate})"
                ),
            ]
        )
    return lines


def build_summary_html(session_summary: dict[str, Any]) -> str:
    lines = build_session_summary_lines(session_summary)
    return (
        '<pre style="margin-top:12px; padding:12px 14px; background:#fafafa; color:#2f2f2f; '
        'border:1px solid #ddd; border-radius:8px; line-height:1.45;">'
        + escape("\n".join(lines))
        + "</pre>"
    )


def format_optional(value: Any, digits: int = 2) -> str:
    if value is None:
        return ""
    return f"{value:.{digits}f}"


def build_table_html(rows: list[dict[str, Any]]) -> str:
    table_rows: list[str] = []
    for row in rows:
        table_rows.append(
            "<tr>"
            f"<td>{row.get('phase', '')}</td>"
            f"<td>{row.get('sanity_kind', '')}</td>"
            f"<td>{row.get('session_trial_number', '')}</td>"
            f"<td>{row.get('attempt_number', '')}</td>"
            f"<td>{row.get('requested_soa_ms', '')}</td>"
            f"<td>{row.get('lead_led', '')}</td>"
            f"<td>{row.get('flash_rgb_level', '')}</td>"
            f"<td>{row.get('expected_response', '')}</td>"
            f"<td>{row.get('response', '')}</td>"
            f"<td>{row.get('sanity_correct', '')}</td>"
            f"<td>{row.get('button', '')}</td>"
            f"<td>{'' if row.get('rt_ms') is None else int(row['rt_ms'])}</td>"
            f"<td>{format_optional(row.get('threshold50_ms'))}</td>"
            f"<td>{format_optional(row.get('ci50_width_ms'))}</td>"
            "</tr>"
        )

    return (
        "<div style='max-height:420px; overflow:auto; border:1px solid #ddd; border-radius:8px;'>"
        "<table style='border-collapse:collapse; width:100%; font-family:Arial, sans-serif; font-size:13px;'>"
        "<thead style='position:sticky; top:0; background:#f6f6f6;'>"
        "<tr>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Phase</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Kind</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Session #</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Trial</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Interval (ms)</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Lead LED</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Flash level</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Expected</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Response</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Correct</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>Button</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>RT (ms)</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>T50 (ms)</th>"
        "<th style='padding:6px; border-bottom:1px solid #ddd;'>CI95 (ms)</th>"
        "</tr></thead><tbody>"
        + "".join(table_rows)
        + "</tbody></table></div>"
    )


def build_report_fragment(
    session_dir: Path,
    session_summary: dict[str, Any],
    rows: list[dict[str, Any]],
    *,
    include_plotlyjs: bool | str = "cdn",
) -> str:
    fig = build_session_figure(session_dir, session_summary, rows)
    plot_html = fig.to_html(
        full_html=False,
        include_plotlyjs=include_plotlyjs,
        config={"displaylogo": False, "responsive": True},
    )
    return (
        "<div style='font-family:Arial, sans-serif;'>"
        + plot_html
        + build_summary_html(session_summary)
        + "<br><b>Sequential patient responses:</b>"
        + build_table_html(rows)
        + "</div>"
    )


def export_plot_html(session_dir: Path) -> Path:
    session_summary, rows = load_session(session_dir)
    fig = build_session_figure(session_dir, session_summary, rows)
    path = session_dir / "session_plot.html"
    fig.write_html(
        str(path),
        full_html=True,
        include_plotlyjs=True,
        config={"displaylogo": False, "responsive": True},
    )
    return path


def export_report_html(session_dir: Path) -> Path:
    session_summary, rows = load_session(session_dir)
    fragment = build_report_fragment(
        session_dir,
        session_summary,
        rows,
        include_plotlyjs=True,
    )
    html = (
        "<!DOCTYPE html><html><head><meta charset='utf-8'>"
        f"<title>TDT session report - {escape(session_dir.name)}</title>"
        "</head><body style='margin:24px; background:#ffffff;'>"
        + fragment
        + "</body></html>"
    )
    path = session_dir / "session_report.html"
    path.write_text(html, encoding="utf-8")
    return path


def flatten_summary_row(session_dir: Path, session_summary: dict[str, Any]) -> dict[str, Any]:
    summary_block = session_summary["summary"]
    estimates = summary_block["estimates"]
    reliability = summary_block["reliability"]
    stop_checks = summary_block["early_stop"]["checks"]
    timing = session_summary.get("session_timing", {})
    sanity = summary_block.get("sanity_checks", {})
    zero_control = sanity.get("zero_control", {})
    easy_control = sanity.get("easy_control", {})

    return {
        "session_name": session_dir.name,
        "trials_completed": summary_block["completed_trials"],
        "trial_limit": session_summary["config"]["trial_count"],
        "sanity_trials_inserted": session_summary.get("sanity_trials_completed", 0),
        "flash_rgb_level": session_summary["config"].get("flash_rgb_level"),
        "threshold50_ms": estimates.get("threshold50_ms"),
        "threshold75_ms": estimates.get("threshold75_ms"),
        "jnd_ms": estimates.get("jnd_ms"),
        "guess_rate": estimates.get("guess_rate", estimates.get("lapse_rate")),
        "lapse_rate": estimates.get("lapse_rate"),
        "ci95_low_ms": reliability.get("threshold50_ci95_low_ms"),
        "ci95_high_ms": reliability.get("threshold50_ci95_high_ms"),
        "ci95_width_ms": reliability.get("threshold50_ci95_width_ms"),
        "ci75_low_ms": reliability.get("threshold75_ci95_low_ms"),
        "ci75_high_ms": reliability.get("threshold75_ci95_high_ms"),
        "ci75_width_ms": reliability.get("threshold75_ci95_width_ms"),
        "posterior_mass_within_precision_window": reliability.get("threshold50_mass_within_precision_window"),
        "normalized_entropy": reliability.get("mean_entropy_normalized"),
        "early_stop_active": summary_block["early_stop"]["active"],
        "early_stop_trials": stop_checks.get("enough_trials"),
        "early_stop_ci50": stop_checks.get("threshold50"),
        "early_stop_entropy": stop_checks.get("entropy"),
        "early_stop_sanity": stop_checks.get("sanity"),
        "total_duration_s": timing.get("total_duration_s"),
        "total_duration_display": timing.get("total_duration_display"),
        "practice_duration_s": timing.get("practice_duration_s"),
        "practice_duration_display": timing.get("practice_duration_display"),
        "adaptive_duration_s": timing.get("adaptive_duration_s"),
        "adaptive_duration_display": timing.get("adaptive_duration_display"),
        "zero_control_correct": zero_control.get("correct_trials"),
        "zero_control_valid": zero_control.get("valid_trials"),
        "zero_control_rate": zero_control.get("correct_rate"),
        "easy_control_correct": easy_control.get("correct_trials"),
        "easy_control_valid": easy_control.get("valid_trials"),
        "easy_control_rate": easy_control.get("correct_rate"),
        "overall_sanity_correct": sanity.get("correct_trials_total"),
        "overall_sanity_valid": sanity.get("valid_trials_total"),
        "overall_sanity_rate": sanity.get("correct_rate_total"),
    }


def _autosize_worksheet(worksheet) -> None:
    if get_column_letter is None:
        return
    for column_cells in worksheet.columns:
        values = [cell.value for cell in column_cells if cell.value is not None]
        if not values:
            continue
        max_length = max(len(str(value)) for value in values)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 40)


def export_summary_workbook(session_dir: Path) -> Path:
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Install openpyxl to export Excel files.")

    session_summary, rows = load_session(session_dir)
    flat = flatten_summary_row(session_dir, session_summary)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"

    header_fill = PatternFill(fill_type="solid", fgColor="E9EEF7")
    header_font = Font(bold=True)

    summary_sheet.append(["Field", "Value"])
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for key, value in flat.items():
        summary_sheet.append([key, value])
    summary_sheet.freeze_panes = "A2"
    _autosize_worksheet(summary_sheet)

    row_sheet = workbook.create_sheet("Summary row")
    row_headers = list(flat.keys())
    row_sheet.append(row_headers)
    row_sheet.append([flat[key] for key in row_headers])
    for cell in row_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    row_sheet.freeze_panes = "A2"
    row_sheet.auto_filter.ref = row_sheet.dimensions
    _autosize_worksheet(row_sheet)

    trials_sheet = workbook.create_sheet("Trials")
    if rows:
        fieldnames = list(rows[0].keys())
        trials_sheet.append(fieldnames)
        for cell in trials_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            trials_sheet.append([row.get(field) for field in fieldnames])
        trials_sheet.freeze_panes = "A2"
        trials_sheet.auto_filter.ref = trials_sheet.dimensions
        _autosize_worksheet(trials_sheet)

    path = session_dir / "summary.xlsx"
    workbook.save(path)
    return path



def build_notebook_report_html(session_dir: Path) -> str:
    session_summary, rows = load_session(session_dir)
    return build_report_fragment(
        session_dir,
        session_summary,
        rows,
        include_plotlyjs="cdn",
    )


def update_notebook_report_output(notebook_path: Path, session_dir: Path) -> Path:
    export_report_html(session_dir)
    notebook_html = build_notebook_report_html(session_dir)

    notebook = json.loads(notebook_path.read_text(encoding="utf-8"))
    code_cell_indexes = [index for index, cell in enumerate(notebook["cells"]) if cell.get("cell_type") == "code"]
    if not code_cell_indexes:
        raise RuntimeError("Notebook does not contain a code cell to update.")

    target_index = code_cell_indexes[-1]
    notebook["cells"][target_index]["outputs"] = [
        {
            "name": "stdout",
            "output_type": "stream",
            "text": [
                f"Loaded session: {session_dir.name}\n",
                "Rendered directly in the notebook output\n",
            ],
        },
        {
            "output_type": "display_data",
            "metadata": {},
            "data": {"text/html": [notebook_html]},
        },
    ]
    notebook["cells"][target_index]["execution_count"] = None
    notebook_path.write_text(json.dumps(notebook, ensure_ascii=False, indent=1), encoding="utf-8")
    return notebook_path


def export_session_artifacts(session_dir: Path, notebook_path: Path | None = None) -> dict[str, str]:
    exported: dict[str, str] = {}
    exported["session_plot_html"] = str(export_plot_html(session_dir))
    exported["session_report_html"] = str(export_report_html(session_dir))
    try:
        exported["summary_xlsx"] = str(export_summary_workbook(session_dir))
    except RuntimeError as exc:
        exported["summary_xlsx_error"] = str(exc)
    if notebook_path is not None and notebook_path.exists():
        exported["notebook_updated"] = str(update_notebook_report_output(notebook_path, session_dir))
    return exported
